# -*- coding: utf-8 -*-
"""
铁路客运站商务候车室管理系统 - Flask后端应用
"""

import os
import json
import uuid
import shutil
import zipfile
import io
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory, render_template
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Protection, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XlImage

from config import (
    STATIONS_FILE, HIERARCHY_FILE, PHOTOS_DIR,
    MAX_CONTENT_LENGTH, ALLOWED_PHOTO_EXTENSIONS
)

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH
CORS(app)

# ============================================================
# 下拉选项配置
# ============================================================
DROPDOWN_OPTIONS = {
    "所属路局": ["哈尔滨铁路局", "沈阳铁路局", "北京铁路局", "太原铁路局", "呼和浩特铁路局",
                "郑州铁路局", "武汉铁路局", "西安铁路局", "济南铁路局", "上海铁路局",
                "南昌铁路局", "广州铁路(集团)公司", "南宁铁路局", "成都铁路局",
                "昆明铁路局", "兰州铁路局", "乌鲁木齐铁路局", "青藏铁路公司"],
    "车站类别": ["一类", "二类", "三类", "四类"],
    "是否为其他功能用房改造": ["是", "否"],
    "安检设备设施设置情况": ["无", "1套"],
    "检验票设备设施设置情况": ["门式闸机", "柱式闸机", "手持检票终端", "无"],
    "是否有独立的站外直接进入通道": ["是", "否"],
    "如何进入其他站台": ["通过候车大厅", "站内专用通道", "通过天桥连廊", "无"],
    "是否有进入公共候车大厅通道": ["是", "否"],
    "站外汽车停靠方式": ["停靠落客平台", "停靠停车场"],
    "是否有独立停车区": ["是", "否"],
    "是否有接待台": ["是", "否"],
    "是否有操作间": ["是", "否"],
    "房屋资产办理情况": [
        "①已签订协议，交铁路集团公司/控股合资铁路公司无偿使用",
        "②已签订协议，无偿移交铁路集团公司/控股合资铁路公司",
        "③已签订协议，交参股合资铁路/地方铁路公司无偿使用",
        "④已签订协议，无偿移交参股合资铁路/地方铁路公司",
        "⑤不移交"
    ],
    "功能调整情况": [
        "①调整为重点旅客、母婴、军人候车等服务用房",
        "②调整为商务座候车区",
        '③调整为\"礼遇行\"经营用房',
        "④调整为生产生活用房"
    ],
    "是否有独立卫生间": ["是", "否"],
    # 商务座候车区
    "类型_商务座": ["高铁站", "高普混", "普速站"],
    "建设出资主体": ["国铁", "铁路局", "地方政府全称", "投资企业全称"],
    "装修出资主体_商务座": ["国铁", "铁路局", "地方政府全称", "投资企业全称"],
    "运营模式": ["自营", "业务外包"],
    "是否有独立进站通道": ["是", "否"],
    "是否有独立安检通道": ["是", "否"],
    "是否有独立检票验证通道": ["是", "否"],
    "是否有独立泊车落客区": ["是", "否"],
    # 商业候车室
    "类型_商业": ["高铁站", "高普混", "普速站"],
    "建设投资主体": ["国铁", "铁路局", "地方政府全称", "投资企业全称"],
    "装修出资主体_商业": ["国铁", "铁路局", "地方政府全称", "投资企业全称"],
    "营业状态": ["营业", "停业"],
    "是否有冠名": ["是", "否"],
    "功能调整_商业": [
        "①调整为重点旅客、母婴、军人候车等服务用房",
        "②调整为商务座候车区",
        '③调整为\"礼遇行\"经营用房',
        "④调整为生产生活用房"
    ],
}

# ============================================================
# 字段定义
# ============================================================

# 模块1：原贵宾候车室 - 公共属性字段（绿色底色）
# 注意：模板中字段顺序为 前4列公共 + 18列休息厅 + 后18列公共
VIP_PUBLIC_FIELDS_FIRST4 = [
    "所属路局", "所属站段", "车站名称", "车站类别",
]

VIP_PUBLIC_FIELDS_LAST18 = [
    "名称", "建成时间", "是否为其他功能用房改造",
    "安检设备设施设置情况", "检验票设备设施设置情况",
    "是否有独立的站外直接进入通道", "如何进入其他站台",
    "是否有进入公共候车大厅通道", "站外汽车停靠方式",
    "是否有独立停车区", "停车位数", "营业时间",
    "专职服务人员", "联系电话", "是否有接待台",
    "卫生间设置数量", "是否有操作间", "车站平面图"
]

# 完整公共属性列表（保持兼容）
VIP_PUBLIC_FIELDS = VIP_PUBLIC_FIELDS_FIRST4 + VIP_PUBLIC_FIELDS_LAST18

# 模块1：原贵宾候车室 - 休息厅属性字段（白色底色）
VIP_LOUNGE_FIELDS = [
    "位置", "休息厅名称", "隶属站房的权属主体单位名称",
    "面积（单位：㎡）", "建设投资主体情况(非施工单位）",
    "房屋资产办理情况", "房屋资产移交公司名称", "房屋资产移交时间",
    "装修出资主体", "装饰资产出资主体", "装修装饰资产明细",
    "装修移交时间", "装饰移交情况", "装饰（家具）移交时间",
    "功能调整情况", "功能调整日期", "是否有独立卫生间", "整改后照片"
]

# 模块1：完整长名称（用于Excel第一行表头）
VIP_PUBLIC_FULL_NAMES_FIRST4 = [
    "所属路局", "所属站段", "车站名称", "车站类别",
]

VIP_PUBLIC_FULL_NAMES_LAST18 = [
    "原贵宾候车室名称", "原贵宾候车室建成时间", "是否为其他功能用房改造",
    "原贵宾候车室安检设备设施设置情况", "原贵宾候车室检验票设备设施设置情况",
    "是否有独立的站外直接进入原贵宾候车室通道", "原贵宾候车室如何进入其他站台",
    "原贵宾候车室是否有进入公共候车大厅通道", "站外汽车停靠方式",
    "原贵宾候车室是否有独立停车区", "原贵宾候车室停车位数",
    "原贵宾候车室营业时间", "原贵宾候车室专职服务人员",
    "原贵宾候车室联系电话", "原贵宾候车室是否有接待台",
    "原贵宾候车室卫生间设置数量", "原贵宾候车室是否有操作间", "车站平面图"
]

VIP_PUBLIC_FULL_NAMES = VIP_PUBLIC_FULL_NAMES_FIRST4 + VIP_PUBLIC_FULL_NAMES_LAST18

VIP_LOUNGE_FULL_NAMES = [
    "原贵宾候车室位置", "原贵宾候车室休息厅名称",
    "原贵宾候车室休息厅隶属站房的权属主体单位名称",
    "原贵宾候车室休息厅面积（单位：㎡）",
    "原贵宾候车室休息厅建设投资主体情况(非施工单位）",
    "原贵宾候车室休息厅房屋资产办理情况",
    "原贵宾候车室休息厅房屋资产移交公司名称",
    "原贵宾候车室休息厅房屋资产移交时间",
    "原贵宾候车室休息厅装修出资主体",
    "原贵宾候车室休息厅装饰资产出资主体",
    "原贵宾候车室休息厅装修装饰资产明细",
    "原贵宾候车室休息厅装修移交时间",
    "原贵宾候车室休息厅装饰移交情况",
    "原贵宾候车室休息厅装饰（家具）移交时间",
    "原贵宾候车室休息厅功能调整情况",
    "原贵宾候车室休息厅功能调整日期",
    "原贵宾候车室休息厅是否有独立卫生间",
    "原贵宾候车室休息厅整改后照片"
]

# 模块2：商务座候车区字段
BUSINESS_CLASS_FIELDS = [
    "所属路局", "所属站段", "车站名称", "车站类别",
    "类型", "设置位置", "建设出资主体（非施工单位）", "建成时间",
    "面积（单位：㎡）", "装修出资主体", "运营模式",
    "委外经营企业名称", "是否有独立进站通道",
    "是否有独立安检通道", "是否有独立检票验证通道",
    "是否有独立泊车落客区", "内设休息间数量", "内设休息间面积"
]

BUSINESS_CLASS_FULL_NAMES = [
    "所属路局", "所属站段", "车站名称", "车站类别",
    "类型", "设置位置", "建设出资主体（非施工单位）", "建成时间",
    "面积（单位：㎡）", "装修出资主体", "运营模式",
    "委外经营企业名称", "是否有独立进站通道",
    "是否有独立安检通道", "是否有独立检票验证通道",
    "是否有独立泊车落客区", "内设休息间数量", "内设休息间面积"
]

# 模块3：商业候车室字段
COMMERCIAL_ROOM_FIELDS = [
    "所属路局", "所属站段", "车站名称", "车站类别",
    "类型", "设置位置", "建设投资主体（非施工单位）", "建成时间",
    "面积（单位：㎡）", "装修出资主体", "经营企业名称",
    "营业状态", "是否有冠名", "冠名单位", "合同到期时间", "功能调整"
]

COMMERCIAL_ROOM_FULL_NAMES = [
    "所属路局", "所属站段", "车站名称", "车站类别",
    "类型", "设置位置", "建设投资主体（非施工单位）", "建成时间",
    "面积（单位：㎡）", "装修出资主体", "经营企业名称",
    "营业状态", "是否有冠名", "冠名单位", "合同到期时间", "功能调整"
]

# 填报说明（第三行，按简化字段名索引）
VIP_FIELD_DESCRIPTIONS = {
    "所属路局": "填报说明：\n手动填写。",
    "所属站段": "填报说明：\n手动填写。",
    "车站名称": "填报说明：\n手动填写。\n①统计范围：一是站房建设设计时为贵宾候车室。二是比照贵宾候车室使用的场所。",
    "车站类别": "填报说明：\n下拉选项，共4项。\n一类\n二类\n三类\n四类",
    "名称": "填报说明：\n①手动填写。\n②与F列贵宾室名称相同，不写休息厅名称即可",
    "建成时间": "填报说明：\n格式示例：200806",
    "是否为其他功能用房改造": "填报说明：\n①下拉选项，2种\n②需核实该贵宾候车室建设期用途。由其他功能用房改造的填\"是\"；在建设期用途为贵宾候车室的，填\"否\"。",
    "安检设备设施设置情况": "填报说明：\n下拉式选项，2种\n①无\n②1套",
    "检验票设备设施设置情况": "填报说明：\n下拉式选项，4种。\n①门式闸机\n②柱式闸机\n③手持检票终端\n④无",
    "是否有独立的站外直接进入通道": "填报：\n下拉式选项，2种。\n①是\n②否",
    "如何进入其他站台": "填报说明：\n下拉式选项，4种。\n①通过候车大厅\n②站内专用通道\n③通过天桥连廊\n④无",
    "是否有进入公共候车大厅通道": "填报说明：\n下拉式选项，2种。\n①是\n②否",
    "站外汽车停靠方式": "填报说明：\n下拉式选项，2种。\n①停靠落客平台\n②停靠停车场",
    "是否有独立停车区": "填报说明：\n下拉式选项，2种\n①是\n②否",
    "停车位数": "填报说明：\n直接填写阿拉伯数字",
    "营业时间": "填写格式XX:XX-XX:XX,24小时营业填写24小时。\n应与候车室营业时间相同。",
    "专职服务人员": "填报说明：\n直接填写阿拉伯数字",
    "联系电话": "填写格式：地方区号+号码",
    "是否有接待台": "填报说明\n下拉式选项，2种。\n①是\n②否",
    "卫生间设置数量": "填报说明：\n直接填写阿拉伯数字。",
    "是否有操作间": "填报说明\n下拉式选项，2种。\n①是\n②否",
    "车站平面图": "注明对应商务候车室位置，底图可结合消防疏散示意图，注明进站流线，每一张图片仅单独标注一处商务候车室位置，用以备案。",
    "位置": "填报说明：\n手动填写。\n①设计时为贵宾室或比照贵宾室使用的场所（区域内一个或多个休息厅按同一位置填写）。\n②一个车站有多处贵宾室的，分别进行填报。\n③贵宾室包括但不限于站房内（外）、基本站台、商服夹层等。\n④填报格式为站名+站房内/外+楼层+方位，如属独栋，需后缀（独栋）。",
    "休息厅名称": "填报说明：\n手动填写。\n①按所属贵宾候车室名称+休息厅名称填报。\n②一处贵宾候车室有多个休息厅的，分别填报。\n③公共休息厅按一个休息厅填报。",
    "隶属站房的权属主体单位名称": "填报说明：\n手动填写。\n①非车站主体站房，而是贵宾室归属站房的权属（因站房建设时间不同，可能会导致贵宾室权属不同的问题）。\n②填写格式\nXX局集团公司;\n控股合资铁路公司-XX公司;\n参股合资铁路-XX公司;\n地方铁路公司-XX公司。\n如有其他，以立项或可研批复为准，据实填写。",
    "面积（单位：㎡）": "填报说明：\n手动填写。\n①面积测量范围仅为休息区域使用面积（不含卫生间、操作间、通廊等）。",
    "建设投资主体情况(非施工单位）": "填报说明：\n手动填写。\n①如建设投资主体和站房权属主体一致，应与G列填写内容一致。\n②如建设投资主体和权属主体不一致，填写具体投资企业全称。",
    "房屋资产办理情况": "填报说明：\n下拉选项，5项。\n①已签订协议，交铁路集团公司/控股合资铁路公司无偿使用\n②已签订协议，无偿移交铁路集团公司/控股合资铁路公司。\n③已签订协议，交参股合资铁路/地方铁路公司无偿使用。\n④已签订协议，无偿移交参股合资铁路/地方铁路公司。\n⑤不移交。",
    "房屋资产移交公司名称": "填写说明：\n手动填写。\n填写对应J列①-④的无偿使用或无偿移交的公司名称。",
    "房屋资产移交时间": "填报说明：\n手动填写。\n①J列显示\"不移交\"，填写\"——\"。\n②无偿移交和无偿使用的，手动填写办结日期；待签协议的，要填写拟办结时间（日期控制在2026年7月31前完成）。\n③办结时间以协议、会议纪要等纸质说明为准。\n④时间格式20260228",
    "装修出资主体": "填报说明：\n手动填写。\n①格式同G列保持一致。",
    "装饰资产出资主体": "填报说明：\n手动填写。\n①格式同G列保持一致。",
    "装修装饰资产明细": "填报说明。\n手动填写。\n①提供主要装修、装饰（家具）资产明细，包含品名、数量。\n②装修是指墙砖、地砖、吊顶等不可拆卸的、不可移动的与房屋建筑融为一体的资产。\n③装饰是指沙发、茶几、屏风、名人字画、花瓶等可搬动的资产。",
    "装修移交时间": "填报说明：\n手动填写\n①此列只填写签订协议时间。时间为：已签订协议，装修资产由投资方按无偿使用交贵宾室所属站房权属主体处理。\n②时间格式20260228",
    "装饰移交情况": "填报说明：\n手动填写。按照下列格式。\n①装饰资产由投资方全部回收。\n②装饰资产由投资方部分回收。回收物品为XX、XXX......（品名、数量）。\n③已签订协议，装饰资产由投资方按无偿使用交贵宾室所属站房权属主体处理。\n④不移交。\n\n需对应留存相关文件、证明。",
    "装饰（家具）移交时间": "填报说明：\n手动填写。\n①Q列显示\"不移交\"，填写\"——\"。\n②无偿使用的、投资方回收或部分回收的，手动填写办结或拟办结时间（日期控制在2026年7月31前完成）。\n③涉及无偿使用的，办结时间以协议、会议纪要等纸质说明为准。\n④时间格式20260228",
    "功能调整情况": "填报说明：\n下拉式选项，共4项。\n\n①调整为重点旅客、母婴、军人候车等服务用房\n②调整为商务座候车区\n③调整为\"礼遇行\"经营用房\n④调整为生产生活用房",
    "功能调整日期": "填报说明：\n①按照实际选择调整日期或拟调整日期（日期控制在2026年11月30前完成）。\n②时间格式20260228",
    "是否有独立卫生间": "填报说明：\n①下拉式选项，共2项。\n②判断标准为休息厅内是否有独立卫生间,而非原贵宾室共用卫生间。\n③选项为是或否",
    "整改后照片": "填报说明：\n门头照片、厅内广角照片至少各1张（要求能显示整改后整体状况）",
}

BC_FIELD_DESCRIPTIONS = {
    "所属路局": "下拉选择",
    "所属站段": "下拉选择（根据路局联动）",
    "车站名称": "下拉选择（根据站段联动），支持搜索",
    "车站类别": "下拉选项，共4项",
    "类型": "下拉选择：高铁站、高普混、普速站",
    "设置位置": "手动填写",
    "建设出资主体（非施工单位）": "下拉选择，共4项",
    "建成时间": "手动填写。时间格式20260228",
    "面积（单位：㎡）": "手动填写数值",
    "装修出资主体": "下拉选择，共4项",
    "运营模式": "下拉选择：自营、业务外包",
    "委外经营企业名称": "运营模式为业务外包时必填",
    "是否有独立进站通道": "下拉选择，共2项",
    "是否有独立安检通道": "下拉选择，共2项",
    "是否有独立检票验证通道": "下拉选择，共2项",
    "是否有独立泊车落客区": "下拉选择，共2项",
    "内设休息间数量": "手动填写数值",
    "内设休息间面积": "手动填写数值",
}

CR_FIELD_DESCRIPTIONS = {
    "所属路局": "下拉选择",
    "所属站段": "下拉选择（根据路局联动）",
    "车站名称": "下拉选择（根据站段联动），支持搜索",
    "车站类别": "下拉选项，共4项",
    "类型": "下拉选择：高铁站、高普混、普速站",
    "设置位置": "手动填写",
    "建设投资主体（非施工单位）": "下拉选择，共4项",
    "建成时间": "手动填写。时间格式20260228",
    "面积（单位：㎡）": "手动填写数值",
    "装修出资主体": "下拉选择，共4项",
    "经营企业名称": "手动填写",
    "营业状态": "下拉选择：营业、停业",
    "是否有冠名": "下拉选择，共2项",
    "冠名单位": "手动填写",
    "合同到期时间": "手动填写",
    "功能调整": "下拉式选项，共4项",
}

# VIP公共属性中需要同步的字段（不含前4个组合键字段和车站平面图）
VIP_PUBLIC_SYNC_FIELDS = [
    "名称", "建成时间", "是否为其他功能用房改造",
    "安检设备设施设置情况", "检验票设备设施设置情况",
    "是否有独立的站外直接进入通道", "如何进入其他站台",
    "是否有进入公共候车大厅通道", "站外汽车停靠方式",
    "是否有独立停车区", "停车位数", "营业时间",
    "专职服务人员", "联系电话", "是否有接待台",
    "卫生间设置数量", "是否有操作间"
]


def _get_vip_public_key(record):
    """获取VIP公共属性组合键（所属路局+所属站段+车站名称+位置）"""
    return (
        record.get('所属路局', ''),
        record.get('所属站段', ''),
        record.get('车站名称', ''),
        record.get('位置', '')
    )


def _sync_vip_public_attributes(data, new_record):
    """同步VIP公共属性：如果存在相同组合键的记录，则自动填充/同步公共属性
    
    规则：
    1. 新增时：如果已有相同组合键的记录，自动填充公共属性
    2. 更新时：修改公共属性后，同步到所有相同组合键的记录
    """
    key = _get_vip_public_key(new_record)
    if not all(key):
        return data
    
    rooms = data.get('vip_rooms', [])
    matching = [r for r in rooms if _get_vip_public_key(r) == key]
    
    if matching:
        # 取第一条匹配记录的公共属性作为基准
        base = matching[0]
        for field in VIP_PUBLIC_SYNC_FIELDS:
            new_val = new_record.get(field, '')
            if new_val:
                # 新记录提供了该字段的值，同步到所有匹配记录
                for r in matching:
                    r[field] = new_val
            else:
                # 新记录未提供该字段的值，从已有记录自动填充
                new_record[field] = base.get(field, '')
    
    return data


# VIP字段到下拉选项key的映射
VIP_FIELD_DROPDOWN_MAP = {
    "所属路局": "所属路局",
    "车站类别": "车站类别",
    "是否为其他功能用房改造": "是否为其他功能用房改造",
    "安检设备设施设置情况": "安检设备设施设置情况",
    "检验票设备设施设置情况": "检验票设备设施设置情况",
    "是否有独立的站外直接进入通道": "是否有独立的站外直接进入通道",
    "如何进入其他站台": "如何进入其他站台",
    "是否有进入公共候车大厅通道": "是否有进入公共候车大厅通道",
    "站外汽车停靠方式": "站外汽车停靠方式",
    "是否有独立停车区": "是否有独立停车区",
    "是否有接待台": "是否有接待台",
    "是否有操作间": "是否有操作间",
    "房屋资产办理情况": "房屋资产办理情况",
    "功能调整情况": "功能调整情况",
    "是否有独立卫生间": "是否有独立卫生间",
}

# 商务座字段到下拉选项key的映射
BC_FIELD_DROPDOWN_MAP = {
    "所属路局": "所属路局",
    "车站类别": "车站类别",
    "类型": "类型_商务座",
    "建设出资主体（非施工单位）": "建设出资主体",
    "装修出资主体": "装修出资主体_商务座",
    "运营模式": "运营模式",
    "是否有独立进站通道": "是否有独立进站通道",
    "是否有独立安检通道": "是否有独立安检通道",
    "是否有独立检票验证通道": "是否有独立检票验证通道",
    "是否有独立泊车落客区": "是否有独立泊车落客区",
}

# 商业候车室字段到下拉选项key的映射
CR_FIELD_DROPDOWN_MAP = {
    "所属路局": "所属路局",
    "车站类别": "车站类别",
    "类型": "类型_商业",
    "建设投资主体（非施工单位）": "建设投资主体",
    "装修出资主体": "装修出资主体_商业",
    "营业状态": "营业状态",
    "是否有冠名": "是否有冠名",
    "功能调整": "功能调整_商业",
}


# ============================================================
# 数据读写工具
# ============================================================

def load_data():
    """加载所有数据"""
    if not os.path.exists(STATIONS_FILE):
        data = {"stations": [], "vip_rooms": [], "business_class_areas": [], "commercial_rooms": []}
        save_data(data)
        return data
    try:
        with open(STATIONS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {"stations": [], "vip_rooms": [], "business_class_areas": [], "commercial_rooms": []}


def save_data(data):
    """保存所有数据"""
    with open(STATIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_hierarchy():
    """加载层级数据"""
    if not os.path.exists(HIERARCHY_FILE):
        return {}
    try:
        with open(HIERARCHY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return {}


def generate_id():
    """生成唯一ID"""
    return str(uuid.uuid4())


def allowed_file(filename):
    """检查文件扩展名是否合法"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_PHOTO_EXTENSIONS


# ============================================================
# 1. 层级数据 API
# ============================================================

@app.route('/api/hierarchy', methods=['GET'])
def get_hierarchy():
    """返回完整层级数据"""
    hierarchy = load_hierarchy()
    return jsonify({"success": True, "data": hierarchy})


@app.route('/api/hierarchy/bureaus', methods=['GET'])
def get_bureaus():
    """返回路局列表"""
    hierarchy = load_hierarchy()
    bureaus = list(hierarchy.keys())
    return jsonify({"success": True, "data": bureaus})


@app.route('/api/hierarchy/sections/<bureau>', methods=['GET'])
def get_sections(bureau):
    """返回指定路局下的站段列表"""
    hierarchy = load_hierarchy()
    bureau = bureau.replace('+', ' ')
    if bureau in hierarchy:
        sections = list(hierarchy[bureau].keys())
        return jsonify({"success": True, "data": sections})
    return jsonify({"success": False, "message": "未找到指定路局"}), 404


@app.route('/api/hierarchy/stations/<bureau>/<section>', methods=['GET'])
def get_stations_by_hierarchy(bureau, section):
    """返回指定站段下的车站列表"""
    hierarchy = load_hierarchy()
    bureau = bureau.replace('+', ' ')
    section = section.replace('+', ' ')
    if bureau in hierarchy and section in hierarchy[bureau]:
        stations = hierarchy[bureau][section]
        return jsonify({"success": True, "data": stations})
    return jsonify({"success": False, "message": "未找到指定路局或站段"}), 404


# ============================================================
# 2. 车站数据管理 API
# ============================================================

@app.route('/api/stations', methods=['GET'])
def get_stations():
    """获取所有车站数据，支持筛选"""
    data = load_data()
    stations = data.get('stations', [])

    bureau = request.args.get('bureau', '').strip()
    section = request.args.get('section', '').strip()
    station_name = request.args.get('station_name', '').strip()

    if bureau:
        stations = [s for s in stations if s.get('所属路局') == bureau]
    if section:
        stations = [s for s in stations if s.get('所属站段') == section]
    if station_name:
        stations = [s for s in stations if s.get('车站名称') == station_name]

    return jsonify({"success": True, "data": stations, "total": len(stations)})


@app.route('/api/stations/<station_id>', methods=['GET'])
def get_station(station_id):
    """获取单个车站详情"""
    data = load_data()
    stations = data.get('stations', [])
    station = next((s for s in stations if s.get('id') == station_id), None)
    if station:
        return jsonify({"success": True, "data": station})
    return jsonify({"success": False, "message": "未找到该车站"}), 404


@app.route('/api/stations', methods=['POST'])
def create_station():
    """新增车站"""
    data = load_data()
    station = request.get_json()
    if not station:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    station['id'] = generate_id()
    station['created_at'] = datetime.now().isoformat()
    station['updated_at'] = datetime.now().isoformat()

    data.setdefault('stations', []).append(station)
    save_data(data)
    return jsonify({"success": True, "data": station}), 201


@app.route('/api/stations/<station_id>', methods=['PUT'])
def update_station(station_id):
    """更新车站"""
    data = load_data()
    stations = data.get('stations', [])
    idx = next((i for i, s in enumerate(stations) if s.get('id') == station_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该车站"}), 404

    update_data = request.get_json()
    if not update_data:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    stations[idx].update(update_data)
    stations[idx]['id'] = station_id  # 保持ID不变
    stations[idx]['updated_at'] = datetime.now().isoformat()
    save_data(data)
    return jsonify({"success": True, "data": stations[idx]})


@app.route('/api/stations/<station_id>', methods=['DELETE'])
def delete_station(station_id):
    """删除车站"""
    data = load_data()
    stations = data.get('stations', [])
    idx = next((i for i, s in enumerate(stations) if s.get('id') == station_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该车站"}), 404

    stations.pop(idx)
    save_data(data)
    return jsonify({"success": True, "message": "删除成功"})


# ============================================================
# 3. 原贵宾候车室 CRUD
# ============================================================

@app.route('/api/vip-rooms', methods=['GET'])
def get_vip_rooms():
    """获取所有原贵宾候车室数据"""
    data = load_data()
    rooms = data.get('vip_rooms', [])

    bureau = request.args.get('bureau', '').strip()
    section = request.args.get('section', '').strip()
    station_name = request.args.get('station_name', '').strip()

    if bureau:
        rooms = [r for r in rooms if r.get('所属路局') == bureau]
    if section:
        rooms = [r for r in rooms if r.get('所属站段') == section]
    if station_name:
        rooms = [r for r in rooms if r.get('车站名称') == station_name]

    return jsonify({"success": True, "data": rooms, "total": len(rooms)})


@app.route('/api/vip-rooms/<room_id>', methods=['GET'])
def get_vip_room(room_id):
    """获取单个原贵宾候车室数据"""
    data = load_data()
    rooms = data.get('vip_rooms', [])
    room = next((r for r in rooms if r.get('id') == room_id), None)
    if room:
        return jsonify({"success": True, "data": room})
    return jsonify({"success": False, "message": "未找到该记录"}), 404


@app.route('/api/vip-rooms/match-public', methods=['GET'])
def match_vip_public():
    """根据所属路局+所属站段+车站名称+位置查找匹配的公共属性"""
    bureau = request.args.get('bureau', '').strip()
    section = request.args.get('section', '').strip()
    station = request.args.get('station', '').strip()
    location = request.args.get('location', '').strip()
    exclude_id = request.args.get('exclude_id', '').strip()  # 编辑时排除当前记录

    if not all([bureau, section, station, location]):
        return jsonify({"success": True, "data": None, "message": "请提供完整的四字段组合"})

    data = load_data()
    rooms = data.get('vip_rooms', [])
    key = (bureau, section, station, location)
    
    for r in rooms:
        if exclude_id and r.get('id') == exclude_id:
            continue  # 跳过当前记录
        if _get_vip_public_key(r) == key:
            # 返回公共属性
            public_data = {}
            for field in VIP_PUBLIC_SYNC_FIELDS:
                public_data[field] = r.get(field, '')
            return jsonify({"success": True, "data": public_data})

    return jsonify({"success": True, "data": None})


@app.route('/api/vip-rooms', methods=['POST'])
def create_vip_room():
    """新增原贵宾候车室（自动同步公共属性）"""
    data = load_data()
    room = request.get_json()
    if not room:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    room['id'] = generate_id()
    room['created_at'] = datetime.now().isoformat()
    room['updated_at'] = datetime.now().isoformat()

    # 同步公共属性：如果已有相同组合键的记录，自动填充公共属性
    data = _sync_vip_public_attributes(data, room)

    data.setdefault('vip_rooms', []).append(room)
    save_data(data)
    return jsonify({"success": True, "data": room}), 201


@app.route('/api/vip-rooms/<room_id>', methods=['PUT'])
def update_vip_room(room_id):
    """更新原贵宾候车室（自动同步公共属性到所有匹配记录）"""
    data = load_data()
    rooms = data.get('vip_rooms', [])
    idx = next((i for i, r in enumerate(rooms) if r.get('id') == room_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该记录"}), 404

    update_data = request.get_json()
    if not update_data:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    rooms[idx].update(update_data)
    rooms[idx]['id'] = room_id
    rooms[idx]['updated_at'] = datetime.now().isoformat()

    # 同步公共属性：修改公共属性后，同步到所有相同组合键的记录
    data = _sync_vip_public_attributes(data, rooms[idx])

    save_data(data)
    return jsonify({"success": True, "data": rooms[idx]})


@app.route('/api/vip-rooms/<room_id>', methods=['DELETE'])
def delete_vip_room(room_id):
    """删除原贵宾候车室"""
    data = load_data()
    rooms = data.get('vip_rooms', [])
    idx = next((i for i, r in enumerate(rooms) if r.get('id') == room_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该记录"}), 404

    rooms.pop(idx)
    save_data(data)

    # 删除关联照片
    photo_dir = os.path.join(PHOTOS_DIR, room_id)
    if os.path.exists(photo_dir):
        shutil.rmtree(photo_dir)

    return jsonify({"success": True, "message": "删除成功"})


# ============================================================
# 4. 商务座候车区 CRUD
# ============================================================

@app.route('/api/business-class-areas', methods=['GET'])
def get_business_class_areas():
    """获取所有商务座候车区数据"""
    data = load_data()
    areas = data.get('business_class_areas', [])

    bureau = request.args.get('bureau', '').strip()
    section = request.args.get('section', '').strip()
    station_name = request.args.get('station_name', '').strip()

    if bureau:
        areas = [a for a in areas if a.get('所属路局') == bureau]
    if section:
        areas = [a for a in areas if a.get('所属站段') == section]
    if station_name:
        areas = [a for a in areas if a.get('车站名称') == station_name]

    return jsonify({"success": True, "data": areas, "total": len(areas)})


@app.route('/api/business-class-areas', methods=['POST'])
def create_business_class_area():
    """新增商务座候车区"""
    data = load_data()
    area = request.get_json()
    if not area:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    area['id'] = generate_id()
    area['created_at'] = datetime.now().isoformat()
    area['updated_at'] = datetime.now().isoformat()

    data.setdefault('business_class_areas', []).append(area)
    save_data(data)
    return jsonify({"success": True, "data": area}), 201


@app.route('/api/business-class-areas/<area_id>', methods=['GET'])
def get_business_class_area(area_id):
    """获取单个商务座候车区数据"""
    data = load_data()
    areas = data.get('business_class_areas', [])
    area = next((a for a in areas if a.get('id') == area_id), None)
    if area:
        return jsonify({"success": True, "data": area})
    return jsonify({"success": False, "message": "未找到该记录"}), 404


@app.route('/api/business-class-areas/<area_id>', methods=['PUT'])
def update_business_class_area(area_id):
    data = load_data()
    areas = data.get('business_class_areas', [])
    idx = next((i for i, a in enumerate(areas) if a.get('id') == area_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该记录"}), 404

    update_data = request.get_json()
    if not update_data:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    areas[idx].update(update_data)
    areas[idx]['id'] = area_id
    areas[idx]['updated_at'] = datetime.now().isoformat()
    save_data(data)
    return jsonify({"success": True, "data": areas[idx]})


@app.route('/api/business-class-areas/<area_id>', methods=['DELETE'])
def delete_business_class_area(area_id):
    """删除商务座候车区"""
    data = load_data()
    areas = data.get('business_class_areas', [])
    idx = next((i for i, a in enumerate(areas) if a.get('id') == area_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该记录"}), 404

    areas.pop(idx)
    save_data(data)

    photo_dir = os.path.join(PHOTOS_DIR, area_id)
    if os.path.exists(photo_dir):
        shutil.rmtree(photo_dir)

    return jsonify({"success": True, "message": "删除成功"})


# ============================================================
# 5. 商业候车室 CRUD
# ============================================================

@app.route('/api/commercial-rooms', methods=['GET'])
def get_commercial_rooms():
    """获取所有商业候车室数据"""
    data = load_data()
    rooms = data.get('commercial_rooms', [])

    bureau = request.args.get('bureau', '').strip()
    section = request.args.get('section', '').strip()
    station_name = request.args.get('station_name', '').strip()

    if bureau:
        rooms = [r for r in rooms if r.get('所属路局') == bureau]
    if section:
        rooms = [r for r in rooms if r.get('所属站段') == section]
    if station_name:
        rooms = [r for r in rooms if r.get('车站名称') == station_name]

    return jsonify({"success": True, "data": rooms, "total": len(rooms)})


@app.route('/api/commercial-rooms', methods=['POST'])
def create_commercial_room():
    """新增商业候车室"""
    data = load_data()
    room = request.get_json()
    if not room:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    room['id'] = generate_id()
    room['created_at'] = datetime.now().isoformat()
    room['updated_at'] = datetime.now().isoformat()

    data.setdefault('commercial_rooms', []).append(room)
    save_data(data)
    return jsonify({"success": True, "data": room}), 201


@app.route('/api/commercial-rooms/<room_id>', methods=['GET'])
def get_commercial_room(room_id):
    """获取单个商业候车室数据"""
    data = load_data()
    rooms = data.get('commercial_rooms', [])
    room = next((r for r in rooms if r.get('id') == room_id), None)
    if room:
        return jsonify({"success": True, "data": room})
    return jsonify({"success": False, "message": "未找到该记录"}), 404


@app.route('/api/commercial-rooms/<room_id>', methods=['PUT'])
def update_commercial_room(room_id):
    """更新商业候车室"""
    data = load_data()
    rooms = data.get('commercial_rooms', [])
    idx = next((i for i, r in enumerate(rooms) if r.get('id') == room_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该记录"}), 404

    update_data = request.get_json()
    if not update_data:
        return jsonify({"success": False, "message": "请求数据为空"}), 400

    rooms[idx].update(update_data)
    rooms[idx]['id'] = room_id
    rooms[idx]['updated_at'] = datetime.now().isoformat()
    save_data(data)
    return jsonify({"success": True, "data": rooms[idx]})


@app.route('/api/commercial-rooms/<room_id>', methods=['DELETE'])
def delete_commercial_room(room_id):
    """删除商业候车室"""
    data = load_data()
    rooms = data.get('commercial_rooms', [])
    idx = next((i for i, r in enumerate(rooms) if r.get('id') == room_id), None)
    if idx is None:
        return jsonify({"success": False, "message": "未找到该记录"}), 404

    rooms.pop(idx)
    save_data(data)

    photo_dir = os.path.join(PHOTOS_DIR, room_id)
    if os.path.exists(photo_dir):
        shutil.rmtree(photo_dir)

    return jsonify({"success": True, "message": "删除成功"})


# ============================================================
# 6. 下拉选项 API
# ============================================================

@app.route('/api/dropdown-options', methods=['GET'])
def get_dropdown_options():
    """返回所有下拉选项配置"""
    return jsonify({"success": True, "data": DROPDOWN_OPTIONS})


# ============================================================
# 7. 照片管理 API
# ============================================================

@app.route('/api/photos/upload', methods=['POST'])
def upload_photos():
    """上传照片（支持多张）"""
    if 'photos' not in request.files and 'file' not in request.files:
        return jsonify({"success": False, "message": "未找到上传文件"}), 400

    record_id = request.form.get('record_id', '')
    if not record_id:
        return jsonify({"success": False, "message": "缺少record_id参数"}), 400

    # 获取文件列表
    files = request.files.getlist('photos')
    if not files:
        files = request.files.getlist('file')
    if not files:
        return jsonify({"success": False, "message": "未找到上传文件"}), 400

    record_dir = os.path.join(PHOTOS_DIR, record_id)
    os.makedirs(record_dir, exist_ok=True)

    uploaded = []
    for file in files:
        if file and file.filename and allowed_file(file.filename):
            ext = file.filename.rsplit('.', 1)[1].lower()
            photo_id = generate_id()
            filename = f"{photo_id}.{ext}"
            filepath = os.path.join(record_dir, filename)
            file.save(filepath)
            uploaded.append({
                "photo_id": photo_id,
                "filename": filename,
                "record_id": record_id,
                "original_name": file.filename,
                "url": f"/api/photos/file/{record_id}/{filename}"
            })

    return jsonify({"success": True, "data": uploaded, "count": len(uploaded)})


@app.route('/api/photos/<record_id>', methods=['GET'])
def get_photos(record_id):
    """获取记录的所有照片"""
    record_dir = os.path.join(PHOTOS_DIR, record_id)
    if not os.path.exists(record_dir):
        return jsonify({"success": True, "data": [], "count": 0})

    photos = []
    for filename in os.listdir(record_dir):
        filepath = os.path.join(record_dir, filename)
        if os.path.isfile(filepath) and allowed_file(filename):
            photo_id = filename.rsplit('.', 1)[0]
            photos.append({
                "photo_id": photo_id,
                "filename": filename,
                "record_id": record_id,
                "url": f"/api/photos/file/{record_id}/{filename}"
            })

    return jsonify({"success": True, "data": photos, "count": len(photos)})


@app.route('/api/photos/file/<record_id>/<filename>', methods=['GET'])
def serve_photo(record_id, filename):
    """提供照片文件访问"""
    record_dir = os.path.join(PHOTOS_DIR, record_id)
    safe_filename = os.path.basename(filename)  # 防止路径遍历
    if os.path.exists(os.path.join(record_dir, safe_filename)):
        return send_from_directory(record_dir, safe_filename)
    return jsonify({"success": False, "message": "文件不存在"}), 404


@app.route('/api/photos/download/<record_id>', methods=['GET'])
def download_photos_zip(record_id):
    """打包下载记录的所有照片（ZIP格式）"""
    record_dir = os.path.join(PHOTOS_DIR, record_id)
    if not os.path.exists(record_dir):
        return jsonify({"success": False, "message": "未找到照片记录"}), 404

    photos = [f for f in os.listdir(record_dir)
              if os.path.isfile(os.path.join(record_dir, f)) and allowed_file(f)]

    if not photos:
        return jsonify({"success": False, "message": "该记录没有照片"}), 404

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for photo in photos:
            filepath = os.path.join(record_dir, photo)
            zipf.write(filepath, photo)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'photos_{record_id}.zip'
    )


@app.route('/api/photos/<photo_id>', methods=['DELETE'])
def delete_photo(photo_id):
    """删除单张照片"""
    # 在所有记录目录中查找该照片
    if os.path.exists(PHOTOS_DIR):
        for record_dir_name in os.listdir(PHOTOS_DIR):
            record_dir = os.path.join(PHOTOS_DIR, record_dir_name)
            if os.path.isdir(record_dir):
                for filename in os.listdir(record_dir):
                    if filename.startswith(photo_id + '.'):
                        filepath = os.path.join(record_dir, filename)
                        os.remove(filepath)
                        return jsonify({"success": True, "message": "照片已删除"})

    return jsonify({"success": False, "message": "未找到该照片"}), 404


# ============================================================
# 8. 模板下载 API
# ============================================================

def _build_sheet_vip(ws, option_ranges=None):
    """构建原贵宾候车室Sheet"""
    # 样式定义
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    desc_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    header_font = Font(bold=True, size=11)
    desc_font = Font(size=9, italic=True, color='666666')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    locked_protection = Protection(locked=True)

    # 字段顺序：前4列公共 + 18列休息厅 + 后18列公共（与新模板一致）
    all_short = VIP_PUBLIC_FIELDS_FIRST4 + VIP_LOUNGE_FIELDS + VIP_PUBLIC_FIELDS_LAST18
    all_full = VIP_PUBLIC_FULL_NAMES_FIRST4 + VIP_LOUNGE_FULL_NAMES + VIP_PUBLIC_FULL_NAMES_LAST18
    total_cols = len(all_short)

    # 判断某列是否为公共属性（绿色）：前4列或后18列
    lounge_start = len(VIP_PUBLIC_FIELDS_FIRST4) + 1  # 第5列开始是休息厅
    lounge_end = lounge_start + len(VIP_LOUNGE_FIELDS) - 1  # 第22列结束

    # 第一行：完整长名称
    for col_idx, full_name in enumerate(all_full, 1):
        cell = ws.cell(row=1, column=col_idx, value=full_name)
        if col_idx < lounge_start or col_idx > lounge_end:
            cell.fill = green_fill
        else:
            cell.fill = white_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    # 第二行：简化短名称
    for col_idx, short_name in enumerate(all_short, 1):
        cell = ws.cell(row=2, column=col_idx, value=short_name)
        if col_idx < lounge_start or col_idx > lounge_end:
            cell.fill = green_fill
        else:
            cell.fill = white_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    # 隐藏第二行
    ws.row_dimensions[2].hidden = True

    # 第三行：填报说明
    for col_idx, short_name in enumerate(all_short, 1):
        desc = VIP_FIELD_DESCRIPTIONS.get(short_name, '')
        cell = ws.cell(row=3, column=col_idx, value=desc)
        cell.fill = desc_fill
        cell.font = desc_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    # 设置行高
    ws.row_dimensions[3].height = 40

    # 设置列宽
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    # 添加数据验证（下拉选项），从第4行开始，使用隐藏Sheet引用
    for col_idx, short_name in enumerate(all_short, 1):
        if short_name in VIP_FIELD_DROPDOWN_MAP:
            option_key = VIP_FIELD_DROPDOWN_MAP[short_name]
            if option_ranges and option_key in option_ranges:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(
                    type="list",
                    formula1=option_ranges[option_key],
                    allow_blank=True
                )
                dv.error = f"请从下拉列表中选择{short_name}"
                dv.errorTitle = "输入错误"
                dv.prompt = f"请选择{short_name}"
                dv.promptTitle = short_name
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}4:{col_letter}1000')

    # === 级联下拉：所属路局 -> 所属站段 -> 车站名称 ===
    # 所属路局(列A)：引用命名区域 _BureauList
    dv_bureau = DataValidation(type="list", formula1="_BureauList", allow_blank=True)
    dv_bureau.prompt = "请先选择所属路局"
    dv_bureau.promptTitle = "所属路局"
    ws.add_data_validation(dv_bureau)
    dv_bureau.add('A4:A1000')

    # 所属站段(列B)：INDIRECT引用路局对应命名区域
    dv_section = DataValidation(type="list", formula1="=INDIRECT(SUBSTITUTE(SUBSTITUTE(A4,\"(\",\"\"),\")\",\"\"))", allow_blank=True)
    dv_section.prompt = "请先选择所属路局，再选择所属站段"
    dv_section.promptTitle = "所属站段"
    dv_section.error = "请先选择所属路局"
    dv_section.errorTitle = "提示"
    ws.add_data_validation(dv_section)
    dv_section.add('B4:B1000')

    # 车站名称(列C)：INDIRECT引用站段对应命名区域（路局简称_站段名）
    dv_station = DataValidation(type="list", formula1='=INDIRECT(LEFT(A4,2)&"_"&B4)', allow_blank=True)
    dv_station.prompt = "请先选择所属站段，再选择车站名称"
    dv_station.promptTitle = "车站名称"
    dv_station.error = "请先选择所属路局和所属站段"
    dv_station.errorTitle = "提示"
    ws.add_data_validation(dv_station)
    dv_station.add('C4:C1000')

    # 将数据填写区域（第4行起）设为未锁定、允许换行
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    unlocked = Protection(locked=False)
    for row_idx in range(4, 1001):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.protection = unlocked
            cell.alignment = data_align
            cell.border = thin_border

    # === 照片列提示：在整改后照片(第22列)和车站平面图(第40列)预填提示文字 ===
    photo_hint_font = Font(size=9, color='0070C0', italic=True)
    photo_hint_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row_idx in range(4, 1001):
        for photo_col in [22, 40]:  # 整改后照片、车站平面图
            cell = ws.cell(row=row_idx, column=photo_col)
            if not cell.value:
                cell.font = photo_hint_font
                cell.alignment = photo_hint_align

    # === 公共属性联动公式 ===
    # 在公共属性列（第23-40列，即W-AN列）添加VLOOKUP公式
    # 当用户填写相同的"所属路局+所属站段+车站名称+位置"时，自动引用第一条匹配记录的公共属性
    # 使用辅助列方式：在每行第41列（AO列，隐藏）拼接组合键，然后用VLOOKUP查找
    public_start_col = 23  # 公共属性起始列（"名称"）
    public_end_col = 40    # 公共属性结束列（"车站平面图"）
    key_col_letter = get_column_letter(41)  # AO列作为辅助组合键列

    # 在辅助列（AO列）添加组合键拼接公式
    for row_idx in range(4, 1001):
        cell = ws.cell(row=row_idx, column=41)
        cell.value = f'=IF(OR(A{row_idx}="",B{row_idx}="",C{row_idx}="",E{row_idx}=""),"",A{row_idx}&"|"&B{row_idx}&"|"&C{row_idx}&"|"&E{row_idx})'
        cell.protection = unlocked

    # 在公共属性列添加VLOOKUP公式（查找上方已有数据）
    for col_idx in range(public_start_col, public_end_col + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(4, 1001):
            cell = ws.cell(row=row_idx, column=col_idx)
            # VLOOKUP: 在AO$4:AO1000中查找当前行的组合键，返回第col_idx列的值
            # MATCH: 找到第一个匹配行的位置
            # IF: 如果组合键为空或没有匹配，则留空
            cell.value = (
                f'=IF(${key_col_letter}{row_idx}="","",'
                f'IFERROR('
                f'INDEX(${col_letter}$4:${col_letter}1000,'
                f'MATCH(${key_col_letter}{row_idx},${key_col_letter}$4:${key_col_letter}1000,0)),'
                f'""'
                f'))'
            )
            cell.protection = unlocked

    # === 时间格式验证（8位数字，如20260228） ===
    time_fields = ['建成时间', '房屋资产移交时间', '装修移交时间', '装饰（家具）移交时间', '功能调整日期']
    for col_idx, short_name in enumerate(all_short, 1):
        if short_name in time_fields:
            col_letter = get_column_letter(col_idx)
            dv_time = DataValidation(
                type="custom",
                formula1=f'=OR(LEN(TRIM({col_letter}4))=0,AND(LEN(TRIM({col_letter}4))=8,ISNUMBER(VALUE({col_letter}4))))',
                allow_blank=True
            )
            dv_time.error = "时间格式错误！请填写8位数字格式，例如：20260228"
            dv_time.errorTitle = "时间格式错误"
            dv_time.prompt = "请填写8位数字时间格式，如20260228"
            dv_time.promptTitle = short_name
            dv_time.showErrorMessage = True
            dv_time.showInputMessage = True
            ws.add_data_validation(dv_time)
            dv_time.add(f'{col_letter}4:{col_letter}1000')

    # 隐藏辅助列（AO列）
    ws.column_dimensions[key_col_letter].hidden = True

    # === 公共属性列条件格式：有公式值时显示浅灰色背景，提示"自动填充，请勿修改" ===
    gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    from openpyxl.formatting.rule import FormulaRule
    for col_idx in range(public_start_col, public_end_col + 1):
        col_letter = get_column_letter(col_idx)
        # 条件：单元格有值（非空）
        rule = FormulaRule(
            formula=[f'AND({col_letter}4<>"",LEN({col_letter}4)>0)'],
            fill=gray_fill
        )
        ws.conditional_formatting.add(f'{col_letter}4:{col_letter}1000', rule)

    # 冻结前两行（第一行表头+第二行隐藏的系统标识）
    ws.freeze_panes = 'A3'

    # 不设置工作表保护，避免openpyxl自动生成密码导致用户无法操作


def _build_sheet_business_class(ws, option_ranges=None):
    """构建商务座候车区Sheet"""
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    desc_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    header_font = Font(bold=True, size=11)
    desc_font = Font(size=9, italic=True, color='666666')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    locked_protection = Protection(locked=True)

    fields = BUSINESS_CLASS_FIELDS
    full_names = BUSINESS_CLASS_FULL_NAMES
    total_cols = len(fields)

    for col_idx, full_name in enumerate(full_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=full_name)
        cell.fill = green_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    for col_idx, short_name in enumerate(fields, 1):
        cell = ws.cell(row=2, column=col_idx, value=short_name)
        cell.fill = green_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    ws.row_dimensions[2].hidden = True

    # 第三行：填报说明
    for col_idx, short_name in enumerate(fields, 1):
        desc = BC_FIELD_DESCRIPTIONS.get(short_name, '')
        cell = ws.cell(row=3, column=col_idx, value=desc)
        cell.fill = desc_fill
        cell.font = desc_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    ws.row_dimensions[3].height = 40

    for col_idx in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for col_idx, short_name in enumerate(fields, 1):
        if short_name in BC_FIELD_DROPDOWN_MAP:
            option_key = BC_FIELD_DROPDOWN_MAP[short_name]
            if option_ranges and option_key in option_ranges:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(
                    type="list",
                    formula1=option_ranges[option_key],
                    allow_blank=True
                )
                dv.error = f"请从下拉列表中选择{short_name}"
                dv.errorTitle = "输入错误"
                dv.prompt = f"请选择{short_name}"
                dv.promptTitle = short_name
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}4:{col_letter}1000')

    # === 级联下拉：所属路局 -> 所属站段 -> 车站名称 ===
    dv_bureau = DataValidation(type="list", formula1="_BureauList", allow_blank=True)
    dv_bureau.prompt = "请先选择所属路局"
    dv_bureau.promptTitle = "所属路局"
    ws.add_data_validation(dv_bureau)
    dv_bureau.add('A4:A1000')

    dv_section = DataValidation(type="list", formula1="=INDIRECT(SUBSTITUTE(SUBSTITUTE(A4,\"(\",\"\"),\")\",\"\"))", allow_blank=True)
    dv_section.prompt = "请先选择所属路局，再选择所属站段"
    dv_section.promptTitle = "所属站段"
    dv_section.error = "请先选择所属路局"
    dv_section.errorTitle = "提示"
    ws.add_data_validation(dv_section)
    dv_section.add('B4:B1000')

    dv_station = DataValidation(type="list", formula1='=INDIRECT(LEFT(A4,2)&"_"&B4)', allow_blank=True)
    dv_station.prompt = "请先选择所属站段，再选择车站名称"
    dv_station.promptTitle = "车站名称"
    dv_station.error = "请先选择所属路局和所属站段"
    dv_station.errorTitle = "提示"
    ws.add_data_validation(dv_station)
    dv_station.add('C4:C1000')

    # 将数据填写区域（第4行起）设为未锁定、允许换行
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    unlocked = Protection(locked=False)
    for row_idx in range(4, 1001):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.protection = unlocked
            cell.alignment = data_align
            cell.border = thin_border

    # === 时间格式验证（建成时间，8位数字） ===
    for col_idx, short_name in enumerate(fields, 1):
        if short_name == '建成时间':
            col_letter = get_column_letter(col_idx)
            dv_time = DataValidation(
                type="custom",
                formula1=f'=OR(LEN(TRIM({col_letter}4))=0,AND(LEN(TRIM({col_letter}4))=8,ISNUMBER(VALUE({col_letter}4))))',
                allow_blank=True
            )
            dv_time.error = "时间格式错误！请填写8位数字格式，例如：20260228"
            dv_time.errorTitle = "时间格式错误"
            dv_time.prompt = "请填写8位数字时间格式，如20260228"
            dv_time.promptTitle = "建成时间"
            dv_time.showErrorMessage = True
            dv_time.showInputMessage = True
            ws.add_data_validation(dv_time)
            dv_time.add(f'{col_letter}4:{col_letter}1000')

    ws.freeze_panes = 'A3'
    # 不设置工作表保护，避免openpyxl自动生成密码导致用户无法操作


def _build_sheet_commercial(ws, option_ranges=None):
    """构建商业候车室Sheet"""
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    desc_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    header_font = Font(bold=True, size=11)
    desc_font = Font(size=9, italic=True, color='666666')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    locked_protection = Protection(locked=True)

    fields = COMMERCIAL_ROOM_FIELDS
    full_names = COMMERCIAL_ROOM_FULL_NAMES
    total_cols = len(fields)

    for col_idx, full_name in enumerate(full_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=full_name)
        cell.fill = green_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    for col_idx, short_name in enumerate(fields, 1):
        cell = ws.cell(row=2, column=col_idx, value=short_name)
        cell.fill = green_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    ws.row_dimensions[2].hidden = True

    # 第三行：填报说明
    for col_idx, short_name in enumerate(fields, 1):
        desc = CR_FIELD_DESCRIPTIONS.get(short_name, '')
        cell = ws.cell(row=3, column=col_idx, value=desc)
        cell.fill = desc_fill
        cell.font = desc_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    ws.row_dimensions[3].height = 40

    for col_idx in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for col_idx, short_name in enumerate(fields, 1):
        if short_name in CR_FIELD_DROPDOWN_MAP:
            option_key = CR_FIELD_DROPDOWN_MAP[short_name]
            if option_ranges and option_key in option_ranges:
                col_letter = get_column_letter(col_idx)
                dv = DataValidation(
                    type="list",
                    formula1=option_ranges[option_key],
                    allow_blank=True
                )
                dv.error = f"请从下拉列表中选择{short_name}"
                dv.errorTitle = "输入错误"
                dv.prompt = f"请选择{short_name}"
                dv.promptTitle = short_name
                ws.add_data_validation(dv)
                dv.add(f'{col_letter}4:{col_letter}1000')

    # === 级联下拉：所属路局 -> 所属站段 -> 车站名称 ===
    dv_bureau = DataValidation(type="list", formula1="_BureauList", allow_blank=True)
    dv_bureau.prompt = "请先选择所属路局"
    dv_bureau.promptTitle = "所属路局"
    ws.add_data_validation(dv_bureau)
    dv_bureau.add('A4:A1000')

    dv_section = DataValidation(type="list", formula1="=INDIRECT(SUBSTITUTE(SUBSTITUTE(A4,\"(\",\"\"),\")\",\"\"))", allow_blank=True)
    dv_section.prompt = "请先选择所属路局，再选择所属站段"
    dv_section.promptTitle = "所属站段"
    dv_section.error = "请先选择所属路局"
    dv_section.errorTitle = "提示"
    ws.add_data_validation(dv_section)
    dv_section.add('B4:B1000')

    dv_station = DataValidation(type="list", formula1='=INDIRECT(LEFT(A4,2)&"_"&B4)', allow_blank=True)
    dv_station.prompt = "请先选择所属站段，再选择车站名称"
    dv_station.promptTitle = "车站名称"
    dv_station.error = "请先选择所属路局和所属站段"
    dv_station.errorTitle = "提示"
    ws.add_data_validation(dv_station)
    dv_station.add('C4:C1000')

    # 将数据填写区域（第4行起）设为未锁定、允许换行
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    unlocked = Protection(locked=False)
    for row_idx in range(4, 1001):
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.protection = unlocked
            cell.alignment = data_align
            cell.border = thin_border

    # === 时间格式验证（建成时间、合同到期时间，8位数字） ===
    for col_idx, short_name in enumerate(fields, 1):
        if short_name in ['建成时间', '合同到期时间']:
            col_letter = get_column_letter(col_idx)
            dv_time = DataValidation(
                type="custom",
                formula1=f'=OR(LEN(TRIM({col_letter}4))=0,AND(LEN(TRIM({col_letter}4))=8,ISNUMBER(VALUE({col_letter}4))))',
                allow_blank=True
            )
            dv_time.error = "时间格式错误！请填写8位数字格式，例如：20260228"
            dv_time.errorTitle = "时间格式错误"
            dv_time.prompt = "请填写8位数字时间格式，如20260228"
            dv_time.promptTitle = short_name
            dv_time.showErrorMessage = True
            dv_time.showInputMessage = True
            ws.add_data_validation(dv_time)
            dv_time.add(f'{col_letter}4:{col_letter}1000')

    ws.freeze_panes = 'A3'
    # 不设置工作表保护，避免openpyxl自动生成密码导致用户无法操作


def _build_dropdown_options_sheet(wb):
    """构建下拉选项Sheet（隐藏），将所有下拉选项写入此Sheet供数据验证引用"""
    ws = wb.create_sheet("下拉选项")
    header_font = Font(bold=True, size=11)

    # 按列写入每个下拉选项组
    col_idx = 1
    option_ranges = {}  # key -> range string (如 "下拉选项!$A$2:$A$20")

    for key, options in DROPDOWN_OPTIONS.items():
        # 第一行写key名称
        ws.cell(row=1, column=col_idx, value=key).font = header_font
        # 从第二行开始写选项
        for i, opt in enumerate(options, 2):
            ws.cell(row=i, column=col_idx, value=opt)
        # 记录范围
        end_row = len(options) + 1
        col_letter = get_column_letter(col_idx)
        option_ranges[key] = f"下拉选项!${col_letter}$2:${col_letter}${end_row}"
        col_idx += 1

    # 隐藏该Sheet
    ws.sheet_state = 'hidden'
    return option_ranges


def _build_hierarchy_sheet(ws, wb=None):
    """构建层级数据参考Sheet（隐藏），同时创建命名区域用于级联下拉
    
    命名区域规则：
    - 路局列表: _BureauList
    - 每个路局下的站段: 路局名（替换特殊字符后）
    - 每个站段下的车站: 路局简称_站段名（替换特殊字符后）
    """
    hierarchy = load_hierarchy()
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')

    # === 第1列：路局列表 ===
    ws.cell(row=1, column=1, value="路局").font = header_font
    ws.cell(row=1, column=1).alignment = center_align
    bureau_names = list(hierarchy.keys())
    for i, bureau in enumerate(bureau_names, 2):
        ws.cell(row=i, column=1, value=bureau)

    # 创建路局列表命名区域
    if wb:
        from openpyxl.workbook.defined_name import DefinedName
        bureau_end_row = len(bureau_names) + 1
        wb.defined_names.add(DefinedName(
            name="_BureauList",
            attr_text=f"层级数据参考!$A$2:$A${bureau_end_row}"
        ))

    # === 第2列起：每个路局一列存放其站段 ===
    col = 2
    for bureau, sections in hierarchy.items():
        ws.cell(row=1, column=col, value=bureau).font = header_font
        ws.cell(row=1, column=col).alignment = center_align
        section_names = list(sections.keys())
        for i, section in enumerate(section_names, 2):
            ws.cell(row=i, column=col, value=section)

        # 创建命名区域：路局名 -> 该路局的站段列表
        if wb:
            safe_name = bureau.replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace(' ', '')
            section_end_row = len(section_names) + 1
            col_letter = get_column_letter(col)
            wb.defined_names.add(DefinedName(
                name=safe_name,
                attr_text=f"层级数据参考!${col_letter}$2:${col_letter}${section_end_row}"
            ))

        col += 1

    # === 后续列：每个站段一列存放其车站 ===
    for bureau, sections in hierarchy.items():
        for section, stations in sections.items():
            ws.cell(row=1, column=col, value=f"{bureau}_{section}").font = header_font
            ws.cell(row=1, column=col).alignment = center_align
            for i, station in enumerate(stations, 2):
                ws.cell(row=i, column=col, value=station)

            # 创建命名区域：路局简称_站段名 -> 该站段的车站列表
            if wb:
                # 路局简称：取前两个字（如"上海铁路局"->"上海"，"广州铁路(集团)公司"->"广州"）
                bureau_short = bureau[:2]
                safe_key = f"{bureau_short}_{section}".replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace(' ', '')
                station_end_row = len(stations) + 1
                col_letter = get_column_letter(col)
                wb.defined_names.add(DefinedName(
                    name=safe_key,
                    attr_text=f"层级数据参考!${col_letter}$2:${col_letter}${station_end_row}"
                ))

            col += 1

    # 隐藏该Sheet
    ws.sheet_state = 'hidden'


@app.route('/api/download-template', methods=['GET'])
def download_template():
    """下载Excel模板"""
    wb = Workbook()

    # Sheet1: 原贵宾候车室
    ws_vip = wb.active
    ws_vip.title = "原贵宾候车室"

    # 先创建下拉选项Sheet（隐藏），获取选项范围引用
    option_ranges = _build_dropdown_options_sheet(wb)

    _build_sheet_vip(ws_vip, option_ranges)

    # Sheet2: 商务座候车区
    ws_bc = wb.create_sheet("商务座候车区")
    _build_sheet_business_class(ws_bc, option_ranges)

    # Sheet3: 商业候车室
    ws_cr = wb.create_sheet("商业候车室")
    _build_sheet_commercial(ws_cr, option_ranges)

    # Sheet4: 层级数据参考（隐藏），同时创建级联下拉命名区域
    ws_h = wb.create_sheet("层级数据参考")
    _build_hierarchy_sheet(ws_h, wb)

    # 保存到内存
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='铁路客运站商务候车室管理系统_导入模板.xlsx'
    )


# ============================================================
# 9. 数据导入 API
# ============================================================

def _read_sheet_data(ws, short_fields, dropdown_map):
    """从工作表读取数据，使用第二行（系统标识行）进行字段匹配"""
    # 读取第二行获取字段映射
    field_map = {}  # col_idx -> short_field_name
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=2, column=col_idx).value
        if cell_value and str(cell_value).strip() in short_fields:
            field_map[col_idx] = str(cell_value).strip()

    # 从第三行开始读取数据
    records = []
    errors = []
    for row_idx in range(4, ws.max_row + 1):
        row_data = {}
        has_data = False
        for col_idx, field_name in field_map.items():
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                cell_value = str(cell_value).strip()
                if cell_value:
                    has_data = True
                    row_data[field_name] = cell_value

        if not has_data:
            continue

        # 校验下拉选项
        for field_name, value in row_data.items():
            if field_name in dropdown_map:
                option_key = dropdown_map[field_name]
                valid_options = DROPDOWN_OPTIONS.get(option_key, [])
                if valid_options and value not in valid_options:
                    errors.append(
                        f"第{row_idx}行：字段'{field_name}'的值'{value}'不在下拉选项中"
                    )

        records.append(row_data)

    return records, errors


@app.route('/api/import', methods=['POST'])
def import_data():
    """导入Excel数据"""
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "未找到上传文件"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({"success": False, "message": "文件名为空"}), 400

    module = request.form.get('module', 'all')  # all, vip, business, commercial

    try:
        from openpyxl import load_workbook
        buffer = io.BytesIO()
        file.save(buffer)
        buffer.seek(0)
        wb = load_workbook(buffer)
    except Exception as e:
        return jsonify({"success": False, "message": f"读取Excel文件失败: {str(e)}"}), 400

    data = load_data()
    total_success = 0
    total_errors = []
    all_errors_detail = []

    # 导入原贵宾候车室
    if module in ('all', 'vip'):
        sheet_name = "原贵宾候车室"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            records, errors = _read_sheet_data(ws, VIP_PUBLIC_FIELDS + VIP_LOUNGE_FIELDS, VIP_FIELD_DROPDOWN_MAP)
            all_errors_detail.extend(errors)

            # 校验公共属性一致性
            public_groups = {}
            for record in records:
                key = (
                    record.get('所属路局', ''),
                    record.get('所属站段', ''),
                    record.get('车站名称', ''),
                    record.get('位置', '')
                )
                if key not in public_groups:
                    public_groups[key] = record
                else:
                    # 检查公共字段是否一致
                    existing = public_groups[key]
                    for pf in VIP_PUBLIC_FIELDS:
                        if pf in ['名称', '建成时间', '是否为其他功能用房改造',
                                  '安检设备设施设置情况', '检验票设备设施设置情况',
                                  '是否有独立的站外直接进入通道', '如何进入其他站台',
                                  '是否有进入公共候车大厅通道', '站外汽车停靠方式',
                                  '是否有独立停车区', '停车位数', '营业时间',
                                  '专职服务人员', '联系电话', '是否有接待台',
                                  '卫生间设置数量', '是否有操作间', '车站平面图']:
                            existing_val = existing.get(pf, '')
                            new_val = record.get(pf, '')
                            if existing_val and new_val and existing_val != new_val:
                                all_errors_detail.append(
                                    f"公共属性不一致：路局={key[0]}, 站段={key[1]}, "
                                    f"车站={key[2]}, 位置={key[3]}, 字段={pf}, "
                                    f"已有值={existing_val}, 新值={new_val}"
                                )

            for record in records:
                record['id'] = generate_id()
                record['created_at'] = datetime.now().isoformat()
                record['updated_at'] = datetime.now().isoformat()
                data.setdefault('vip_rooms', []).append(record)
                total_success += 1

    # 导入商务座候车区
    if module in ('all', 'business'):
        sheet_name = "商务座候车区"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            records, errors = _read_sheet_data(ws, BUSINESS_CLASS_FIELDS, BC_FIELD_DROPDOWN_MAP)
            all_errors_detail.extend(errors)

            for record in records:
                record['id'] = generate_id()
                record['created_at'] = datetime.now().isoformat()
                record['updated_at'] = datetime.now().isoformat()
                data.setdefault('business_class_areas', []).append(record)
                total_success += 1

    # 导入商业候车室
    if module in ('all', 'commercial'):
        sheet_name = "商业候车室"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            records, errors = _read_sheet_data(ws, COMMERCIAL_ROOM_FIELDS, CR_FIELD_DROPDOWN_MAP)
            all_errors_detail.extend(errors)

            for record in records:
                record['id'] = generate_id()
                record['created_at'] = datetime.now().isoformat()
                record['updated_at'] = datetime.now().isoformat()
                data.setdefault('commercial_rooms', []).append(record)
                total_success += 1

    save_data(data)

    return jsonify({
        "success": True,
        "message": f"导入完成",
        "imported": total_success,
        "errors": len(all_errors_detail),
        "error_details": all_errors_detail
    })


# ============================================================
# 10. 数据导出 API
# ============================================================

def _export_records_to_sheet(ws, records, short_fields, full_names, is_vip=False, descriptions=None):
    """将记录导出到工作表"""
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    desc_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    header_font = Font(bold=True, size=11)
    desc_font = Font(size=9, italic=True, color='666666')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    locked_protection = Protection(locked=True)

    total_cols = len(short_fields)

    # 第一行：完整长名称
    for col_idx, full_name in enumerate(full_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=full_name)
        if is_vip and col_idx <= len(VIP_PUBLIC_FIELDS):
            cell.fill = green_fill
        else:
            cell.fill = green_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    # 第二行：简化短名称
    for col_idx, short_name in enumerate(short_fields, 1):
        cell = ws.cell(row=2, column=col_idx, value=short_name)
        if is_vip and col_idx <= len(VIP_PUBLIC_FIELDS):
            cell.fill = green_fill
        else:
            cell.fill = green_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.protection = locked_protection

    ws.row_dimensions[2].hidden = True

    # 第三行：填报说明
    if descriptions:
        for col_idx, short_name in enumerate(short_fields, 1):
            desc = descriptions.get(short_name, '')
            cell = ws.cell(row=3, column=col_idx, value=desc)
            cell.fill = desc_fill
            cell.font = desc_font
            cell.alignment = center_align
            cell.border = thin_border
            cell.protection = locked_protection
        ws.row_dimensions[3].height = 40

    for col_idx in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # 填充数据（从第四行开始）
    for row_idx, record in enumerate(records, 4):
        for col_idx, field_name in enumerate(short_fields, 1):
            value = record.get(field_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border

    # VIP公共属性字段合并单元格
    if is_vip and len(records) > 1:
        # 按路局+站段+车站分组，合并公共属性列
        pass  # 合并逻辑较复杂，此处仅标记

    ws.freeze_panes = 'A3'


@app.route('/api/export', methods=['POST'])
def export_data():
    """导出Excel数据"""
    data = load_data()
    module = request.get_json().get('module', 'all') if request.is_json else 'all'

    wb = Workbook()

    if module in ('all', 'vip'):
        ws_vip = wb.active
        ws_vip.title = "原贵宾候车室"
        records = data.get('vip_rooms', [])
        _export_records_to_sheet(
            ws_vip, records,
            VIP_PUBLIC_FIELDS + VIP_LOUNGE_FIELDS,
            VIP_PUBLIC_FULL_NAMES + VIP_LOUNGE_FULL_NAMES,
            is_vip=True,
            descriptions=VIP_FIELD_DESCRIPTIONS
        )

    if module in ('all', 'business'):
        ws_bc = wb.create_sheet("商务座候车区") if module == 'all' else wb.active
        if module == 'all':
            ws_bc = wb.create_sheet("商务座候车区")
        else:
            ws_bc = wb.active
            ws_bc.title = "商务座候车区"
        records = data.get('business_class_areas', [])
        _export_records_to_sheet(
            ws_bc, records,
            BUSINESS_CLASS_FIELDS,
            BUSINESS_CLASS_FULL_NAMES,
            descriptions=BC_FIELD_DESCRIPTIONS
        )

    if module in ('all', 'commercial'):
        if module == 'all':
            ws_cr = wb.create_sheet("商业候车室")
        else:
            ws_cr = wb.active
            ws_cr.title = "商业候车室"
        records = data.get('commercial_rooms', [])
        _export_records_to_sheet(
            ws_cr, records,
            COMMERCIAL_ROOM_FIELDS,
            COMMERCIAL_ROOM_FULL_NAMES,
            descriptions=CR_FIELD_DESCRIPTIONS
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='铁路客运站商务候车室管理系统_导出数据.xlsx'
    )


# ============================================================
# 页面路由
# ============================================================

@app.route('/', methods=['GET'])
def index():
    """渲染主页面"""
    return render_template('index.html')


# ============================================================
# 健康检查
# ============================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        "success": True,
        "message": "系统运行正常",
        "timestamp": datetime.now().isoformat()
    })


# ============================================================
# 主入口
# ============================================================

if __name__ == '__main__':
    # 确保数据目录存在
    os.makedirs(os.path.dirname(STATIONS_FILE), exist_ok=True)
    os.makedirs(PHOTOS_DIR, exist_ok=True)

    app.run(host='0.0.0.0', port=5000, debug=True)
